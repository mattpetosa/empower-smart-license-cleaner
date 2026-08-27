from pathlib import Path

from openpyxl import load_workbook

from excel import build_workbook
from licenses import parse_pdf, parse_text

SAMPLE = Path(__file__).resolve().parent.parent / "fixtures" / "sample-empower370.pdf"


def test_sample_pdf_rules():
    res = parse_pdf(SAMPLE.read_bytes())
    assert res.installation == "EmpowerDemo"
    assert res.printed.startswith("08/20/2026")
    serials = [l.serial for l in res.licenses]
    # suffix stripped
    assert "M9FJA7135U" in serials and not any(s.endswith("-001") for s in serials)
    # named-user pack sharing the base serial is folded into the base row
    base = [l for l in res.licenses if l.category == "Base License"]
    assert len(base) == 1 and base[0].serial == "R1ZCR5103X" and base[0].qty == 5
    assert serials.count("R1ZCR5103X") == 1
    assert any("Included with base" in r.reason for r in res.removed)
    # nothing duplicated
    keys = [(l.name, l.serial) for l in res.licenses]
    assert len(keys) == len(set(keys))
    assert len(res.licenses) == 19 and not res.unparsed
    cats = {l.category for l in res.licenses}
    assert "Instrument Control (3rd Party)" in cats and "Options" in cats


def test_duplicates_after_suffix_strip():
    text = """Waters Licensing Wizard : X
   [Empower 3 System Control License Pack] System licenses: 2 Serial No: AAA-001
   [Empower 3 System Control License Pack] System licenses: 2 Serial No: AAA-002
   [Empower 3 Agilent LC] Instrument control licenses: 1 Serial No: BBB
   [Empower 3 Agilent LC] Instrument control licenses: 1 Serial No: BBB
   [Empower 3 SystemsQT License] Serial No: CCC-003
"""
    res = parse_text(text)
    assert [l.serial for l in res.licenses] == ["AAA", "BBB", "CCC"]
    assert [r.reason for r in res.removed] == ["Duplicate serial"] * 2


def test_workbook():
    res = parse_pdf(SAMPLE.read_bytes())
    wb = load_workbook(filename=__import__("io").BytesIO(build_workbook(res)))
    assert wb.sheetnames == ["Licenses", "Summary", "Removed"]
    ws = wb["Licenses"]
    assert ws.max_row == 1 + len(res.licenses)
    assert ws["A1"].value == "Category" and ws["E2"].value == "R1ZCR5103X"


def test_csv_exports():
    from excel import build_csv
    res = parse_pdf(SAMPLE.read_bytes())
    simple = build_csv(res, simple=True).split("\r\n")
    assert simple[0] == "Serial_Numbers" and simple[1] == "R1ZCR5103X" and simple[-1] == ""
    assert len(simple) == 1 + len(res.licenses) + 1
    detailed = build_csv(res, simple=False).split("\r\n")
    assert detailed[0].startswith("Category,License,Quantity") and len(detailed) == len(simple)


def test_remove_sqt():
    from licenses import remove_sqt
    res = remove_sqt(parse_pdf(SAMPLE.read_bytes()))
    assert not any("sqt" in l.name.lower() for l in res.licenses)
    assert len(res.licenses) == 13 and sum("SQT removed" in r.reason for r in res.removed) == 6
    assert any("System Suitability" in l.name for l in res.licenses)


def test_default_qty_and_remove_zero():
    from licenses import remove_zero_qty
    text = """Waters Licensing Wizard : X
   [Empower 3 System Suitability] Serial No: A1
   [Empower 3 GPC/SEC Option] Serial No: A2
   [Empower 3 Dissolution] Serial No: A3
   [Empower 3 SystemsQT License] Serial No: A4
   [Empower 3 System Control License Pack] System licenses: 0 Serial No: A5
   [Empower 3 System Control License Pack] System licenses: 2 Serial No: A6
"""
    res = parse_text(text)
    by = {l.serial: l.qty for l in res.licenses}
    assert by == {"A1": 1, "A2": 1, "A3": 1, "A4": None, "A5": 0, "A6": 2}
    remove_zero_qty(res)
    assert sorted(l.serial for l in res.licenses) == ["A1", "A2", "A3", "A4", "A6"]
    assert res.removed[-1].reason.startswith("Zero quantity")


def test_checksum_txt():
    from licenses import parse_upload
    res = parse_upload((SAMPLE.parent / "sample-checksum.txt").read_bytes())
    assert res.company == "Example Pharma Inc" and res.support_id == "EM3SA00000"
    assert res.installation == "EMPOWERSVR"
    base = [l for l in res.licenses if l.category == "Base License"]
    assert len(base) == 1 and base[0].serial == "R2PNC4607S" and base[0].qty == 5
    serials = [l.serial for l in res.licenses]
    assert len(serials) == len(set(serials)) and not any("-00" in s for s in serials)
    assert serials.count("Q2VZ3S848C") == 1  # Shimadzu Control + License(s) same key -> one row
    cats = {l.category for l in res.licenses}
    assert "Instrument Control (3rd Party)" in cats and "System Control" in cats and "User Licenses" in cats
    by = {l.serial: l for l in res.licenses}
    assert by["N9YQB3628E"].qty == 1 and by["M4QHG2190T"].qty == 1  # System Suitability, Dissolution
    assert "sqt" in by["N0TSV0026V"].name.lower()


def test_default_qty_matches_whole_words_only():
    """"sec" used to match "Empower Security Option" as a substring, so a
    license the Wizard deliberately prints without a count arrived in the
    Waters workbook claiming a quantity of 1 that Waters never stated."""
    text = """Waters Licensing Wizard : X
   [Empower Security Option] Serial No: B1
   [Empower 3 Secure Data Option] Serial No: B2
   [Empower 3 GPC/SEC Option] Serial No: B3
   [Empower 3 System Suitability] Serial No: B4
   [Empower 3 Dissolution] Serial No: B5
   [Empower 3 GPC] Serial No: B6
"""
    res = parse_text(text)
    assert {l.serial: l.qty for l in res.licenses} == {
        "B1": None, "B2": None, "B3": 1, "B4": 1, "B5": 1, "B6": 1}


def test_unreadable_checksum_option_lines_are_recorded():
    """The grammar wants two or more spaces before "Serial Number"; a report
    printed with a tab or a single space used to vanish without a trace.
    It is still not parsed, but it now shows up on the Removed sheet
    instead of leaving the workbook quietly short a row."""
    from licenses import parse_checksum_text
    text = (
        "Company Name - Example Pharma Inc\n"
        "Option Properly Installed - 5 Named User License(s)          "
        "Serial Number - W3SAP2033M-001\n"
        "Option Properly Installed - 1 Dissolution License(s)\t"
        "Serial Number - Q1ABC2345D\n"
        "Option Properly Installed - 1 GPC License(s) "
        "Serial Number - Q1ABC2346E\n"
        "CRC 0x8842aa noise line that is not a license at all\n"
    )
    res = parse_checksum_text(text)
    assert [l.serial for l in res.licenses] == ["W3SAP2033M"]
    assert len(res.unparsed) == 2
    assert all("Option Properly Installed" in u for u in res.unparsed)
    assert not any("CRC" in u for u in res.unparsed)


def test_untrusted_cells_cannot_become_formulas():
    """A license name or serial is text out of a Waters report, not something
    we wrote. openpyxl turns a leading "=" into a formula cell outright, and
    Excel evaluates a CSV field starting with =, +, - or @ on import."""
    import io as _io

    from excel import build_csv
    from licenses import License, Result

    res = Result(installation="=HYPERLINK(1)", printed=None, licenses=[
        License(name="=cmd|'/c calc'!A1", serial="@SUM(1)", qty=1,
                qty_label=None, raw="[=evil] Serial No: @SUM(1)",
                category="Options")])
    res.removed.append(__import__("licenses").Removed("-2+3", "Duplicate serial"))

    wb = load_workbook(filename=_io.BytesIO(build_workbook(res)))
    for sheet in wb.sheetnames:
        for row in wb[sheet].iter_rows():
            for c in row:
                assert c.data_type != "f", f"{sheet}!{c.coordinate} is a formula"
    assert wb["Licenses"]["B2"].value == "'=cmd|'/c calc'!A1"

    csv_rows = build_csv(res, simple=True).split("\r\n")
    assert csv_rows[1] == "'@SUM(1)"


def test_summary_category_header_keeps_its_header_font():
    """Column A was bolded row by row afterwards, which replaced the white
    header font on "Category" with plain black — unreadable on the navy fill
    while "Licenses" and "Total Quantity" beside it stayed white."""
    import io as _io

    res = parse_pdf(SAMPLE.read_bytes())
    wb = load_workbook(filename=_io.BytesIO(build_workbook(res)))
    ws = wb["Summary"]
    header = next(r for r in ws.iter_rows() if r[0].value == "Category")
    assert header[0].font.color.rgb == header[1].font.color.rgb
    assert header[0].font.color.rgb.endswith("FFFFFF")


def test_pdf_header_is_found_past_a_byte_order_mark():
    """The spec allows %PDF anywhere in the first 1024 bytes and readers
    follow suit, so a file that picked up a BOM or a banner on its way
    through an email gateway is still a PDF — it used to be handed to the
    text decoder and rejected as "not a Licensing Wizard PDF"."""
    from licenses import parse_upload
    padded = b"\xef\xbb\xbf" + SAMPLE.read_bytes()
    res = parse_upload(padded)
    assert res.installation == "EmpowerDemo" and len(res.licenses) == 19


def test_a_pdf_that_hangs_poppler_is_a_400_not_a_500():
    """subprocess.TimeoutExpired is not a ValueError, so it escaped
    _get_result()'s handler and became a 500 with no message for the user."""
    import subprocess

    import licenses as licenses_mod

    def boom(*a, **kw):
        raise subprocess.TimeoutExpired(cmd="pdftotext", timeout=60)

    real, subprocess.run = subprocess.run, boom
    try:
        with __import__("pytest").raises(ValueError) as e:
            licenses_mod.pdf_to_text(b"%PDF-1.4 whatever")
    finally:
        subprocess.run = real
    assert "too long" in str(e.value)


def test_simple_csv_never_carries_details():
    """The Waters portal upload chokes on anything below the serial list."""
    from excel import build_csv
    res = parse_pdf(SAMPLE.read_bytes())
    rows = build_csv(res, simple=True, details=[("Company", "Acme"), ("Support ID", "123")]).split("\r\n")
    assert rows[0] == "Serial_Numbers"
    assert all(r and " " not in r and ":" not in r for r in rows[1:-1]), rows
    assert rows[-1] == ""
