"""Build the .xlsx workbook from a parsed Result."""
from __future__ import annotations

import io
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from licenses import CATEGORY_ORDER, Result

def inert(value):
    """Stop a cell from being read as a live formula.

    Everything here comes out of a Waters report, and a license name or a
    raw dropped line is text we did not write. openpyxl turns any string
    starting with "=" into a formula cell outright, and Excel and Sheets
    both evaluate a CSV field starting with =, +, - or @ on import - the
    classic CSV-injection path, which on Excel reaches DDE. A leading
    apostrophe is the standard escape that forces the cell to stay text.
    """
    if isinstance(value, str) and value[:1] in ("=", "+", "-", "@"):
        return "'" + value
    return value


HEADER_FILL = PatternFill("solid", fgColor="1F3A5F")
HEADER_FONT = Font(bold=True, color="FFFFFF")
GROUP_FILL = PatternFill("solid", fgColor="E8EEF6")
THIN = Side(style="thin", color="C9D1D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _header(ws, cols):
    ws.append(cols)
    r = ws.max_row
    for c in ws[r]:
        c.fill, c.font, c.border = HEADER_FILL, HEADER_FONT, BORDER
        c.alignment = Alignment(vertical="center")
    ws.freeze_panes = f"A{r + 1}"
    return r


def _autosize(ws, minimum=10, maximum=60):
    widths = {}
    for row in ws.iter_rows():
        for c in row:
            if c.value is not None:
                widths[c.column] = max(widths.get(c.column, 0), len(str(c.value)))
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = max(minimum, min(maximum, w + 2))


def build_workbook(res: Result, details: list[tuple[str, str]] | None = None) -> bytes:
    details = details or []
    wb = Workbook()
    ws = wb.active
    ws.title = "Licenses"
    if details:
        # Company / Support ID / Sold To on the first rows, then a blank line,
        # then the table - so the identifying info is right there when printed.
        for label, value in details:
            ws.append([label, inert(value)])
            ws[ws.max_row][0].font = Font(bold=True)
        ws.append([])
    hdr = _header(ws, ["Category", "License", "Quantity", "Quantity Type", "Serial No"])
    last_cat = None
    for lic in res.licenses:
        ws.append([lic.category, inert(lic.name), lic.qty,
                   inert(lic.qty_label), inert(lic.serial)])
        row = ws[ws.max_row]
        for c in row:
            c.border = BORDER
        if lic.category != last_cat:
            for c in row:
                c.fill = GROUP_FILL
            row[0].font = Font(bold=True)
            last_cat = lic.category
        else:
            row[0].font = Font(color="8A94A6")
        row[2].alignment = Alignment(horizontal="center")
    ws.auto_filter.ref = f"A{hdr}:E{ws.max_row}"
    _autosize(ws)

    summary = wb.create_sheet("Summary")
    for label, value in details:
        summary.append([label, inert(value)])
    summary.append(["Installation", inert(res.installation or "")])
    summary.append(["Date Printed", inert(res.printed or "")])
    summary.append(["Licenses (after cleanup)", len(res.licenses)])
    summary.append(["Lines removed", len(res.removed)])
    summary.append([])
    summary.append(["Category", "Licenses", "Total Quantity"])
    cat_header_row = summary.max_row
    for c in summary[cat_header_row]:
        c.fill, c.font, c.border = HEADER_FILL, HEADER_FONT, BORDER
    counts = Counter(l.category for l in res.licenses)
    qty = Counter()
    for l in res.licenses:
        qty[l.category] += l.qty or 0
    for cat in CATEGORY_ORDER:
        if counts[cat]:
            summary.append([cat, counts[cat], qty[cat]])
    # Skips the category header row: bolding column A unconditionally
    # replaced its white-on-navy header font with plain black bold, so
    # "Category" alone came out unreadable against the dark fill while
    # "Licenses" and "Total Quantity" beside it stayed white.
    for r in summary.iter_rows(min_row=1, max_col=1):
        if r[0].row != cat_header_row:
            r[0].font = Font(bold=True)
    _autosize(summary)

    removed = wb.create_sheet("Removed")
    _header(removed, ["Reason", "Original Line"])
    for r in res.removed:
        removed.append([r.reason, inert(r.raw)])
    for line in res.unparsed:
        removed.append(["Unrecognized line (not a license)", inert(line)])
    _autosize(removed, maximum=110)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_csv(res: Result, simple: bool, details: list[tuple[str, str]] | None = None) -> str:
    """Detailed: same columns as the Licenses sheet. Simple: the Waters
    licensing-portal upload format - a single `Serial_Numbers` column, one
    cleaned serial per line, CRLF, deduplicated in order, nothing else
    (`details` is accepted for signature parity but ignored)."""
    import csv

    buf = io.StringIO()
    w = csv.writer(buf, lineterminator="\r\n")
    if simple:
        w.writerow(["Serial_Numbers"])
        for s in dict.fromkeys(l.serial for l in res.licenses):
            w.writerow([inert(s)])
        # The Waters portal rejects a file with anything below the serial
        # list, so the details (company, support ID...) stay in the filename
        # and in the workbook only - never in this column.
    else:
        w.writerow(["Category", "License", "Quantity", "Quantity Type", "Serial No"])
        for l in res.licenses:
            w.writerow([l.category, inert(l.name),
                        "" if l.qty is None else l.qty,
                        inert(l.qty_label or ""), inert(l.serial)])
    return buf.getvalue()
